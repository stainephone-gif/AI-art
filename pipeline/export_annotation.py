"""Слепая таблица для ручной разметки: reports/annotation_sheet.xlsx.

Состав: все тексты, у которых хотя бы одна v3-модель дала не-UND, плюс случайная выборка UND
в объёме und_sample_fraction от них, стратифицированно по источнику. Слой трудной проблемы
размечается на всех текстах (флаг annotate_hp = 1). Порядок строк перемешан, seed фиксирован.
Меток моделей в таблице нет. Ключ выборки (почему текст попал в подвыборку) сохраняется отдельно
в reports/annotation_sample_key.csv — его нельзя показывать кодировщикам.

python -m pipeline.export_annotation --config config.yaml [--all]
"""

from __future__ import annotations

import argparse
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .agreement import load_v3_theory
from .common import ALL_CLASSES, load_codebook, load_config, load_corpus, reports_dir, resolve

CODERS = ("A", "B")
AUTHOR_TYPES = ["artist", "curator", "joint", "institution"]


def select_sample(corpus: pd.DataFrame, nonund_ids: set, frac: float, seed: int) -> pd.DataFrame:
    """Возвращает corpus с колонками annotate_theory, sample_reason."""
    df = corpus.copy()
    df["sample_reason"] = np.where(df["id"].isin(nonund_ids), "model_non_und", "")
    n_target = int(round(len(nonund_ids) * frac))
    und = df[~df["id"].isin(nonund_ids)]
    rng = np.random.default_rng(seed)
    chosen: List[int] = []
    if n_target > 0 and len(und) > 0:
        # стратификация по источнику пропорционально числу UND-текстов в источнике
        shares = und.groupby("source").size() / len(und)
        alloc = {s: int(np.floor(shares[s] * n_target)) for s in shares.index}
        rest = n_target - sum(alloc.values())
        for s in sorted(shares.index, key=lambda s: -(shares[s] * n_target - alloc[s]))[:rest]:
            alloc[s] += 1
        for s, k in alloc.items():
            pool = und[und["source"] == s]["id"].tolist()
            k = min(k, len(pool))
            chosen += [int(x) for x in rng.choice(pool, size=k, replace=False)] if k else []
    df.loc[df["id"].isin(chosen), "sample_reason"] = "und_sample"
    df["annotate_theory"] = (df["sample_reason"] != "").astype(int)
    df["annotate_hp"] = 1
    return df


def build_sheet(cfg: Dict[str, Any], include_all: bool = False, out_path=None) -> pd.DataFrame:
    seed = int(cfg.get("seed", 0))
    frac = float(cfg.get("annotation", {}).get("und_sample_fraction", 0.3))
    corpus = load_corpus(cfg)
    th = load_v3_theory(cfg)
    if include_all or th.empty:
        if th.empty and not include_all:
            print("v3-прогонов нет: все тексты помечены для разметки рамок (как при --all)")
        nonund = set(corpus["id"])
    else:
        nonund = set(th.loc[(th["status"] == "ok") & (th["primary"] != "UND"), "id"])
    df = select_sample(corpus, nonund, frac, seed)
    key = df[["id", "source", "title", "annotate_theory", "sample_reason"]]

    # слепая таблица, перемешанная
    rng = np.random.default_rng(seed)
    order = rng.permutation(len(df))
    sheet = df.iloc[order][["id", "source", "title", "text", "n_words", "annotate_theory", "annotate_hp"]].reset_index(drop=True)
    for c in CODERS:
        for col in (f"theory_primary_{c}", f"theory_secondary_{c}", f"hp_level_{c}", f"hp_stance_{c}", f"note_{c}"):
            sheet[col] = ""

    author = df[["id", "source", "title", "author"]].copy()
    author["author_type"] = ""

    codebook_lines = load_codebook(cfg).splitlines()
    codebook_df = pd.DataFrame({"codebook": codebook_lines})
    instr = pd.DataFrame({"instruction": [
        "Лист annotation: каждая строка — один текст. Колонки с суффиксом _A заполняет кодировщик A, с суффиксом _B — кодировщик B, независимо и не глядя друг на друга.",
        "annotate_theory = 1: размечать теоретическую рамку (theory_primary, theory_secondary). annotate_hp = 1: размечать слой трудной проблемы (hp_level, hp_stance). hp размечается для всех текстов.",
        f"theory_primary: один из {', '.join(ALL_CLASSES)}. UND — если ни для одного класса нет фрагмента с переносом между областями (см. codebook).",
        "theory_secondary: дополнительные классы через |, только если для каждого есть не менее двух независимых фрагментов; иначе пусто.",
        "hp_level: 0, 1 или 2. hp_stance: denial | attribution | open_question | reframing — только при hp_level ≥ 1.",
        "note: цитаты (дословно), на которых основано решение, и сомнения. Цитаты обязательны для любой не-UND метки и для hp_level ≥ 1.",
        "Лист codebook: определения и критерии исключения — те же, что в промптах моделей.",
        "Лист author_type: проставить тип автора текста: artist | curator | joint | institution.",
        "Не меняйте порядок строк и значения id.",
    ]})

    out_path = out_path or resolve(cfg, cfg["paths"]["annotation_sheet"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        instr.to_excel(xw, sheet_name="instructions", index=False)
        sheet.to_excel(xw, sheet_name="annotation", index=False)
        codebook_df.to_excel(xw, sheet_name="codebook", index=False)
        author.to_excel(xw, sheet_name="author_type", index=False)
        ws = xw.sheets["annotation"]
        cols = list(sheet.columns)
        n = len(sheet) + 1
        dv_class = DataValidation(type="list", formula1='"' + ",".join(ALL_CLASSES) + '"', allow_blank=True)
        dv_level = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
        dv_stance = DataValidation(type="list", formula1='"denial,attribution,open_question,reframing"', allow_blank=True)
        for dv in (dv_class, dv_level, dv_stance):
            ws.add_data_validation(dv)
        for c in CODERS:
            dv_class.add(f"{get_column_letter(cols.index(f'theory_primary_{c}') + 1)}2:{get_column_letter(cols.index(f'theory_primary_{c}') + 1)}{n}")
            dv_level.add(f"{get_column_letter(cols.index(f'hp_level_{c}') + 1)}2:{get_column_letter(cols.index(f'hp_level_{c}') + 1)}{n}")
            dv_stance.add(f"{get_column_letter(cols.index(f'hp_stance_{c}') + 1)}2:{get_column_letter(cols.index(f'hp_stance_{c}') + 1)}{n}")
        ws.column_dimensions[get_column_letter(cols.index("text") + 1)].width = 90
        ws.column_dimensions[get_column_letter(cols.index("title") + 1)].width = 30
        wa = xw.sheets["author_type"]
        dv_author = DataValidation(type="list", formula1='"' + ",".join(AUTHOR_TYPES) + '"', allow_blank=True)
        wa.add_data_validation(dv_author)
        dv_author.add(f"E2:E{len(author) + 1}")
        xw.sheets["codebook"].column_dimensions["A"].width = 140
        xw.sheets["instructions"].column_dimensions["A"].width = 140
    key.to_csv(reports_dir(cfg) / "annotation_sample_key.csv", index=False)
    n_theory = int(df["annotate_theory"].sum())
    print(f"annotation_sheet: {len(sheet)} строк; для разметки рамок {n_theory} "
          f"(не-UND по моделям {len(nonund)}, выборка UND {int((df['sample_reason'] == 'und_sample').sum())}); "
          f"файл {out_path}")
    return sheet


def main(argv: Optional[List[str]] = None) -> None:
    ap = argparse.ArgumentParser(description="Экспорт слепой таблицы для ручной разметки")
    ap.add_argument("--config", default="config.yaml")
    ap.add_argument("--all", action="store_true", help="включить все тексты в подвыборку для разметки рамок")
    ap.add_argument("--out", default=None)
    args = ap.parse_args(argv)
    cfg = load_config(args.config)
    build_sheet(cfg, include_all=args.all, out_path=resolve(cfg, args.out) if args.out else None)


if __name__ == "__main__":
    main()
